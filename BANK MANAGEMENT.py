
try :
	import os
	import random
	from openpyxl import load_workbook,Workbook
	import openpyxl
	from time import sleep
	from datetime import date
	import pyfiglet
except ModuleNotFoundError:
	os.system("python -m pip install colorama openpyxl pyfiglet")


path = "Account Details.xlsx"
#ACC_No= 124300000
#CustomerID=110000000
Today=date.today()
Out="Hii"


def CBA():
        sleep(3)
        os.system("cls")
        os.system("cls")
        logo()
        print(" _____________________________________________________________________________________________________________________________________________________________________________________________")
        print("|\t\t\t\t\t\t\t\t\tCREATE BANK ACCOUNT                                                                                                   |")
        print("|_____________________________________________________________________________________________________________________________________________________________________________________________|\n\n")
        Name=str(input("Name: "))
        Surname=str(input("Surname: "))
        Address=str(input("Address: "))
        SM=int(input("Entry Money: "))
        Adhar=str(input("Do you have aadhar card (Yes/No): "))
        if Adhar=="Yes":
            AdharNo=int(input("Aadhar Number: "))
        else:
            AdharNo=000000
        Pan=str(input("Do you have pan card (Yes/No): "))
        if Pan=="Yes":
            PanNo=int(input("Pan Number: "))
        else:
            PanNo=0000000
        Password=str(input("Password: "))
        print("_______________________________________________________________________________________________________________________________________________________________________________________________\n\n")
        ATM_No1=str(random.randrange(1000,9999))
        ATM_No2=str(random.randrange(1000,9999))
        ATM_No3=str(random.randrange(1000,9999))
        ATM_No4=str(random.randrange(1000,9999))
        ATM_No=ATM_No1 + " " + ATM_No2 + " " + ATM_No3 + " " + ATM_No4

        Today=date.today()

        wb = load_workbook(path)
        ws = wb.active
        m_row = len(ws['a'])
        row_AccountNo=ws.cell(m_row,column=7).value
        row_CustomerID=ws.cell(m_row,column=8).value
        AccountNo= int(row_AccountNo) + 1
        CustomerID= int(row_CustomerID) + 1


        LFULLNAME=Name+" "+Surname
        ws.append([Name+" "+Surname,Address,SM,AdharNo,PanNo,Password,AccountNo,CustomerID,ATM_No,Today])
        wb.save(path)
        openpass(LFULLNAME,CustomerID,AccountNo,Password)
        return Name+" "+Surname,Address,SM,AdharNo,PanNo,Password,AccountNo,CustomerID

        
def openpass(LFULLNAME,LID,LAccountNo,Password):
        wb = load_workbook(filename=path)
        ws = wb.active
        m_row = len(ws['a'])
        for i in range(2,m_row+1):
            row_name=ws.cell(row=i,column=1).value
            row_address=ws.cell(row=i,column=2).value
            row_SM=ws.cell(row=i,column=3).value
            row_Adhar=ws.cell(row=i,column=4).value
            row_Pan=ws.cell(row=i,column=5).value
            row_Password=ws.cell(row=i,column=6).value
            row_AccountNo=ws.cell(row=i,column=7).value 
            row_CustomerID=ws.cell(row=i,column=8).value
            
            if LFULLNAME == row_name and LID == row_CustomerID and LAccountNo == row_AccountNo and Password == row_Password:
                print("_____________________________________________________________________________________________________________________________________________________________________________________________")
                print("Successful!!!! :)")
                print("_____________________________________________________________________________________________________________________________________________________________________________________________")
                try:
                    a=load_workbook(f'{row_CustomerID}_{LFULLNAME}_PASSBOOK (Main).xlsx')
                except:
                    a=openpyxl.load_workbook('PASSBOOK (Main).xlsx')
                    
                sheet = a.active
                sheet['E9']=LFULLNAME
                sheet['E10']=row_address
                sheet['E8']=row_AccountNo
                sheet['E7']=row_CustomerID
                sheet['E11']= date.today()
                sheet['A15']= date.today()
                sheet['B15']="CREDITED"
                sheet['C15']="NA"
                sheet['D15']=row_SM
                sheet['E15']=row_SM
                open=f'{row_CustomerID}_{LFULLNAME}_PASSBOOK (Main).xlsx'
                a.save(open)
                os.system(open)
                sleep(3)
                os.system("cls")
                os.system("cls")

                return "Yes"
            elif LFULLNAME!=row_name and i==m_row:
                print("Not Found !!")
                print("Please Try Again :)")
                sleep(3)
                sleep(2)
                os.system("cls")
                

def Mchoice(Method,LFULLNAME,CustomerID):
        CH=+1
        CREDIT=0
        DEBIT=0
        if Method== 'A':
                        while True:
                            os.system("cls")
                            os.system("cls")
                            sleep(1)
                            logo()
                            print(" _____________________________________________________________________________________________________________________________________________________________________________________________")
                            print("|\t\t\t\t\t\t\t\t\t    CREDIT                                                                                                            |")
                            print("|_____________________________________________________________________________________________________________________________________________________________________________________________|\n\n")
                            CREDIT=int(input("How much amount you want to Credit in your Bank Account: "))
                            Check=str(input("Are you sure you want to Credit {}Rs in your account.(Yes/No): ".format(CREDIT)))
                            if Check=='Yes':
                                    print("Ok. Beacause of your confirmation we are crediting {}Rs in your Bank Account.".format(CREDIT))
                                    print("Wait while we are updating your Bank passbook. :)")
                                    sleep(3)
                                    print("Your Bank Passbook Have Been Update Please Check Your Bank Passbook")
                                    sleep(3)
                                    a=load_workbook(f'{CustomerID}_{LFULLNAME}_PASSBOOK (Main).xlsx')
                                    b=a.active
                                    ws=len(b['E'])
                                    for i in range(2,ws+1):
                                        SM=b.cell(row=i,column=5).value
                                    ADD=SM+CREDIT
                                    b.append([Today,"CREDITED","NA",CREDIT,ADD])
                                    a.save(f'{CustomerID}_{LFULLNAME}_PASSBOOK (Main).xlsx')
                                    sleep(2)
                                    os.system("cls")
                                    break
                            else:
                                    print(" _____________________________________________________________________________________________________________________________________________________________________________________________")
                                    print("\t\t\t\t\t\tOK!! Please write proper balance again...")
                                    print(" _____________________________________________________________________________________________________________________________________________________________________________________________")
                                    sleep(2)
                                    os.system("cls")
                        os.system(f'{CustomerID}_{LFULLNAME}_PASSBOOK (Main).xlsx')
                        
        elif Method== 'B':
                        while True:
                            os.system("cls")
                            os.system("cls")
                            sleep(1)
                            logo()
                            print(" _____________________________________________________________________________________________________________________________________________________________________________________________")
                            print("|\t\t\t\t\t\t\t\t\t    DEBIT                                                                                                             |")
                            print("|_____________________________________________________________________________________________________________________________________________________________________________________________|\n\n")
                            DEBIT=int(input("How much amount you want to Debit in your Bank Account: "))
                            Check=str(input("Are you sure you want to Debit {}Rs in your account.(Yes/No): ".format(DEBIT)))
                            if Check=='Yes':
                                    print("Ok. Beacause of your confirmation we are debiting {}Rs from your Bank Account.".format(DEBIT))
                                    print("Your Bank Passbook will be update within 3 second.")
                                    sleep(3)
                                    print("Your Bank Passbook Have Been Update Please Check Your Bank Passbook")
                                    sleep(3)
                                    a=load_workbook(f'{CustomerID}_{LFULLNAME}_PASSBOOK (Main).xlsx')
                                    b=a.active
                                    ws=len(b['E'])
                                    for i in range(2,ws+1):
                                        SM=b.cell(row=i,column=5).value
                                    ADD=SM-DEBIT
                                    b.append([Today,"DEBITED",DEBIT,"NA",ADD])
                                    a.save(f'{CustomerID}_{LFULLNAME}_PASSBOOK (Main).xlsx')
                                    sleep(1)
                                    os.system("cls")
                                    sleep(1)
                                    break
                            else:
                                    print(" _____________________________________________________________________________________________________________________________________________________________________________________________")
                                    print("\t\t\t\t\t\tOK!! Please write proper balance again...")
                                    print(" _____________________________________________________________________________________________________________________________________________________________________________________________")
                                    sleep(2)
                                    os.system("cls")
                                    sleep(1)
                        os.system(f"{LFULLNAME}_PASSBOOK (Main).xlsx")
        else:
            print("Wrong method!!! \nPlease enter once again..")
            sleep(2)



def logo():
	os.system("pyfiglet --width=150 -f slant Banking Management")




while True:
    os.system("cls")
    os.system("cls")
    logo()
    print(" _____________________________________________________________________________________________________________________________________________________________________________________________")
    print("|\t\t\t                                                                                                                                                                      |")
    print("|\t\t\t\t\t    1 ] Create Bank Account                                                                                                                           |")
    print("|\t\t\t\t\t    2 ] Login Acconut                                                                                                                                 |")
    print("|\t\t\t\t\t    3 ] ATM                                                                                                                                           |")
    print("|\t\t\t\t\t    4 ] Exit                                                                                                                                          |")
    print("|_____________________________________________________________________________________________________________________________________________________________________________________________|")
    Choice=int(input("\n\t\t\t\t            Enter Choice in Number : "))

    if Choice==1:
        Fullname,Address,SM,AdharNo,PanNo,Password,AccountNo,CustomerID=CBA()
        openpass(Fullname,CustomerID,AccountNo,Password)
        os.system("cls")
        os.system("cls")
        sleep(2)
    elif Choice==2:
        while True:
            os.system("cls")
            os.system("cls")
            sleep(2)
            logo()
            print(" _____________________________________________________________________________________________________________________________________________________________________________________________")
            print("|\t\t\t\t\t\t\t\t\t    LOGIN ACCOUNT                                                                                                     |")
            print("|_____________________________________________________________________________________________________________________________________________________________________________________________|\n\n")
            LFULLNAME=str(input("NAME & SURNAME: "))
            CustomerID=int(input("Customer ID: "))
            LAccountNo=int(input("Account No.: "))
            Password=str(input("Password: "))
            print("______________________________________________________________________________________________________________________________________________________________________________________________\n\n")
            if True:
                Out=openpass(LFULLNAME,CustomerID,LAccountNo,Password)
            if Out=="Yes":
                    break
            
        
        print("______________________________________________________________________________________________________________________________________________________________________________________________\n\n")
        print("Select One Option From Given.\n")
        print("A . Credit\n")
        print("B . Debit\n")
        print("______________________________________________________________________________________________________________________________________________________________________________________________\n")
        Method=str(input("Enter which method you want to choose: "))
        Mchoice(Method,LFULLNAME,CustomerID)
        os.system("cls")
        os.system("cls")

    elif Choice==3:
        while True:
            os.system("cls")
            os.system("cls")
            sleep(2)
            logo()
            print(" ______________________________________________________________________________________________________________________________________________________________________________________________")
            print("|\t\t\t\t\t\t\t\t\t    ATM                                                                                                               |")
            print("|_____________________________________________________________________________________________________________________________________________________________________________________________|\n\n")
            LFULLNAME=str(input("NAME & SURNAME: "))
            LID=int(input("Customer ID: "))
            LAccountNo=int(input("Account No.: "))
            Password=str(input("Password: "))
            print(" ______________________________________________________________________________________________________________________________________________________________________________________________")
            wb = load_workbook(filename=path)
            ws = wb.active
            m_row = len(ws['a'])
            for i in range(2,m_row+1):
                row_name=ws.cell(row=i,column=1).value
                row_Password=ws.cell(row=i,column=6).value
                row_AccountNo=ws.cell(row=i,column=7).value 
                row_CustomerID=ws.cell(row=i,column=8).value
                row_ATM_No=ws.cell(row=i,column=9).value
                row_Acc_Open_date=ws.cell(row=i,column=10).value
                if LFULLNAME == row_name and LID == row_CustomerID and LAccountNo == row_AccountNo and Password == row_Password:
                    ATM= openpyxl.load_workbook('BOI ATM CARD.xlsx')
                    ATM_sheet=ATM.active
                    ATM_sheet['B6']=row_ATM_No
                    ATM_sheet['B8']=LFULLNAME
                    ATM_sheet['D8']=row_Acc_Open_date
                    ATMNAME=f'{row_CustomerID}_{LFULLNAME}_BOI ATM CARD (Main).xlsx'
                    ATM.save(ATMNAME)
                    os.system(ATMNAME)
                    Out="Out"
                    print("Your ATM Have Been Created Please Check :)")
                    sleep(2)
                    break
                elif LFULLNAME!=row_name and i==m_row:
                    print(" _____________________________________________________________________________________________________________________________________________________________________________________________")
                    print("|\t\t\t\t\t\t\t\t\t    ERROR                                                                                                             |")
                    print("|_____________________________________________________________________________________________________________________________________________________________________________________________|\n\n")
                    print("Not Found!!!")
                    print("Do you really make you bank account?  ")
                    print(" ______________________________________________________________________________________________________________________________________________________________________________________________")
                    sleep(3)
                    os.system("cls")
                    os.system("cls")
                    break
            if Out=="Out":
                 os.system("cls")
                 os.system("cls")
                 break
        
    elif Choice==4:
        print("Wait 3 second for exit :)")
        print("Thanks For Coming Bank Management App :) ....")
        print("Bye Bye!!! :) ....")
        quit()
    else:
        print("Wrong choice please!!! \nPlease enter once again...")
        sleep(2)
    
    input()
           
